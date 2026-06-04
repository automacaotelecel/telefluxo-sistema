import json
import sys
import traceback
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT_FILENAME = None
OUTPUT_FILENAME = "RECEBIMENTOS STONE TRATADOS.xlsx"

SHEET_BASE_CANDIDATES = ["BASE TRATADA"]
SHEET_EXTRATO_CANDIDATES = ["EXTRATO BANCARIO", "EXTRATO_BANCARIO"]

HOLIDAY_API_URL = "https://date.nager.at/api/v3/PublicHolidays/{year}/BR"
REQUEST_TIMEOUT = 20
HOLIDAY_CACHE_FILENAME = "feriados_cache_br.json"

ADDITIONAL_MANUAL_HOLIDAYS = []
EXTRATO_ONLY_POSITIVE_VALUES = True

# =============================================================================
# PERFORMANCE / EXCEL
# =============================================================================
# A base enviada gera mais de 350 mil previsões após abrir o parcelamento.
# Gravar e FORMATAR todas essas linhas em abas detalhadas deixa o Excel/openpyxl
# extremamente lento e parece que o programa travou.
#
# Por padrão, o arquivo final sai com PAINEL, CONSOLIDADO, RESUMOS,
# CONCILIAÇÃO, EXTRATO e PREVISTO CONCATENADO completos.
# As abas muito pesadas ficam como AMOSTRA das primeiras linhas.
# Se quiser exportar tudo mesmo assim, troque para True, mas pode demorar muitos minutos.
EXPORTAR_ABAS_DETALHADAS_COMPLETAS = False
LIMITE_LINHAS_ABAS_DETALHADAS = 5000


BASE_COL_ALIASES = {
    "data_venda": ["data de venda", "data da venda", "data_venda", "data venda"],
    "valor_liquido": ["valor liquido", "valor líquido", "valor_liq", "vlr liquido", "vlr líquido", "vlr liq"],
    "valor_parcela": ["vlr parcela", "valor parcela", "valor da parcela", "vlr_parcela", "parcela liquida", "parcela líquida"],
    "produto": ["produto", "tipo", "modalidade"],
    "parcelas": ["n de parcelas", "n parcelas", "parcelas", "numero de parcelas", "nro parcelas", "qtde parcelas"],
    "documento": ["documento", "doc", "numero documento", "n documento"],
    "bandeira": ["bandeira", "cartao", "cartão", "bandeira cartao", "bandeira cartão"],
}

EXTRATO_COL_ALIASES = {
    "data_extrato": [
        "data", "data extrato", "data lancamento", "data lançamento", "data do credito", "data do crédito",
        "data recebimento", "data pagamento", "data movimento", "data movimentacao", "data movimentação"
    ],
    "valor_extrato": [
        "valor", "valor recebido", "valor credito", "valor crédito", "credito", "crédito",
        "entrada", "deposito", "depósito", "valor lancamento", "valor lançamento",
        "valor do credito", "valor do crédito", "valor liquido", "valor líquido"
    ],
    "credito_extrato": ["credito", "crédito", "entrada", "valor credito", "valor crédito", "deposito", "depósito"],
    "debito_extrato": ["debito", "débito", "saida", "saída", "valor debito", "valor débito"],
    "descricao": ["descricao", "descrição", "historico", "histórico", "detalhe", "complemento"],
}

WEEKDAY_PT = {
    0: "Segunda-Feira",
    1: "Terça-Feira",
    2: "Quarta-Feira",
    3: "Quinta-Feira",
    4: "Sexta-Feira",
    5: "Sábado",
    6: "Domingo",
}


def normalize_text(text):
    if text is None:
        return ""
    text = str(text).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    for token in ["_", "-", "/", "\\", ".", "(", ")", "[", "]", ":"]:
        text = text.replace(token, " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def get_app_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    if "__file__" in globals():
        return Path(__file__).resolve().parent
    return Path.cwd()


def detect_input_file(base_dir: Path) -> Path:
    if INPUT_FILENAME:
        path = base_dir / INPUT_FILENAME
        if not path.exists():
            raise FileNotFoundError(f"Arquivo configurado não encontrado: {path}")
        return path

    candidates = [
        p for p in base_dir.iterdir()
        if p.is_file()
        and p.name.lower() != OUTPUT_FILENAME.lower()
        and p.suffix.lower() in {".xlsb", ".xlsx", ".xlsm"}
        and not p.name.startswith("~$")
    ]

    if not candidates:
        raise FileNotFoundError("Não encontrei nenhum arquivo .xlsb/.xlsx/.xlsm na mesma pasta do programa.")

    candidates.sort(key=lambda p: (0 if p.suffix.lower() == ".xlsb" else 1, -p.stat().st_mtime))
    return candidates[0]


def get_engine(path: Path):
    if path.suffix.lower() == ".xlsb":
        return "pyxlsb"
    return None


def resolve_sheet_name(xls: pd.ExcelFile, candidates):
    normalized_map = {normalize_text(name): name for name in xls.sheet_names}

    for candidate in candidates:
        key = normalize_text(candidate)
        if key in normalized_map:
            return normalized_map[key]

    for candidate in candidates:
        key = normalize_text(candidate)
        for norm_name, original in normalized_map.items():
            if key == norm_name or key in norm_name or norm_name in key:
                return original

    raise ValueError(f"Não encontrei a aba esperada. Abas disponíveis: {xls.sheet_names}")


def build_normalized_column_map(df: pd.DataFrame):
    return {normalize_text(col): col for col in df.columns}


def find_column(df: pd.DataFrame, aliases, required=True):
    normalized_map = build_normalized_column_map(df)

    for alias in aliases:
        key = normalize_text(alias)
        if key in normalized_map:
            return normalized_map[key]

    for alias in aliases:
        key = normalize_text(alias)
        for norm_col, original_col in normalized_map.items():
            if key in norm_col or norm_col in key:
                return original_col

    if required:
        raise ValueError(f"Não encontrei a coluna. Tente uma destas variações: {aliases}")

    return None


def get_series(df: pd.DataFrame, col_name):
    if col_name is None:
        return pd.Series([None] * len(df), index=df.index)

    obj = df[col_name]
    if isinstance(obj, pd.DataFrame):
        return obj.iloc[:, 0]
    return obj


def parse_money(value):
    if pd.isna(value):
        return np.nan

    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)

    s = str(value).strip()
    if not s:
        return np.nan

    s = s.replace("R$", "").replace("r$", "")
    s = s.replace(" ", "").replace("\u00A0", "")

    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return np.nan


def parse_money_series(series: pd.Series) -> pd.Series:
    return series.apply(parse_money)


def _parse_single_date(value):
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value.normalize()

    if isinstance(value, datetime):
        return pd.Timestamp(value.date())

    if isinstance(value, date):
        return pd.Timestamp(value)

    if isinstance(value, (int, float, np.integer, np.floating)):
        num = float(value)

        if np.isnan(num) or np.isinf(num):
            return pd.NaT

        if 20000 <= num <= 80000:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(num), unit="D")).normalize()

        if 1_000_000_000_000 <= num <= 9_999_999_999_999:
            return pd.to_datetime(int(num), unit="ms", errors="coerce").normalize()

        if 1_000_000_000 <= num <= 9_999_999_999:
            return pd.to_datetime(int(num), unit="s", errors="coerce").normalize()

        if 19_000_000 <= num <= 29_999_999 and float(num).is_integer():
            s = str(int(num))
            return pd.to_datetime(s, format="%Y%m%d", errors="coerce")

        return pd.NaT

    s = str(value).strip()
    if not s:
        return pd.NaT

    s_clean = s.replace(".", "/").replace("-", "/")

    for fmt in [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%Y%m%d",
    ]:
        try:
            return pd.Timestamp(datetime.strptime(s_clean, fmt).date())
        except Exception:
            pass

    if s.isdigit():
        if len(s) == 13:
            return pd.to_datetime(int(s), unit="ms", errors="coerce").normalize()
        if len(s) == 10:
            return pd.to_datetime(int(s), unit="s", errors="coerce").normalize()
        if len(s) == 8 and s.startswith(("19", "20")):
            return pd.to_datetime(s, format="%Y%m%d", errors="coerce")

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return pd.NaT
    return dt.normalize()


def parse_date_series(series: pd.Series) -> pd.Series:
    return series.apply(_parse_single_date)


def parse_int_safe(valor, default=1):
    if pd.isna(valor):
        return default

    s = str(valor).strip()
    if not s:
        return default

    # Quando vier algo como "2/12", usa o número antes da barra.
    if "/" in s:
        s = s.split("/")[0]

    s = s.replace(",", ".")
    try:
        out = int(float(s))
        return out if out > 0 else default
    except Exception:
        return default


def is_credito(produto) -> bool:
    return "CREDITO" in normalize_text(produto)


def is_debito(produto) -> bool:
    return "DEBITO" in normalize_text(produto)


def is_pix(produto) -> bool:
    return "PIX" in normalize_text(produto)


def calculate_rule_schedule(produto, parcela_atual, total_parcelas):
    prod = normalize_text(produto)
    parcela_atual = parse_int_safe(parcela_atual, 1)
    total_parcelas = parse_int_safe(total_parcelas, 1)

    # REGRA CORRETA DO FLUXO STONE USADA NESTE ARQUIVO:
    # - DÉBITO: D+1 + 30 dias corridos, depois joga para o próximo dia útil.
    # - PIX: D+1 corrido, depois joga para o próximo dia útil.
    # - CRÉDITO 1X: D+30 + 30 dias corridos, depois joga para o próximo dia útil.
    # - CRÉDITO PARCELADO: gera uma linha para CADA parcela.
    #   Parcela 1: D+(1x30)+30
    #   Parcela 2: D+(2x30)+30
    #   Parcela 3: D+(3x30)+30 ...
    #
    # O erro anterior era jogar o VALOR LÍQUIDO inteiro em uma única data,
    # usando N DE PARCELAS como se fosse a parcela atual. Agora o código
    # expande a venda em várias linhas e usa VLR PARCELA quando existir.

    if "DEBITO" in prod:
        return 1, 30, "DÉBITO = D+1 + 30 corridos → recebimento no próximo útil"

    if "PIX" in prod:
        return 1, 0, "PIX = D+1 corrido → recebimento no próximo útil"

    if "CREDITO" in prod:
        if total_parcelas <= 1:
            return 30, 30, "CRÉDITO 1X = D+30 + 30 corridos → recebimento no próximo útil"
        return parcela_atual * 30, 30, f"CRÉDITO PARCELADO {parcela_atual}/{total_parcelas} = D+({parcela_atual}x30) + 30 corridos → recebimento no próximo útil"

    raise ValueError(f"Produto não suportado para regra de recebimento: {produto}")

def is_business_day(d: date, holidays: set) -> bool:
    return d.weekday() < 5 and d not in holidays


def move_to_next_business_day(d: date, holidays: set) -> date:
    current = d
    while not is_business_day(current, holidays):
        current += timedelta(days=1)
    return current


def calculate_receipt_flow(sale_date: date, base_calendar_days: int, delay_calendar_days: int, holidays: set, memo=None):
    if memo is None:
        memo = {}

    key = (sale_date.isoformat(), int(base_calendar_days), int(delay_calendar_days))
    if key in memo:
        return memo[key]

    start_count_date = move_to_next_business_day(sale_date, holidays)
    liquidation_candidate = start_count_date + timedelta(days=base_calendar_days)
    liquidation_date = move_to_next_business_day(liquidation_candidate, holidays)
    receipt_candidate = liquidation_date + timedelta(days=delay_calendar_days)
    receipt_date = move_to_next_business_day(receipt_candidate, holidays)

    result = {
        "data_inicio_contagem": start_count_date,
        "data_liquidacao": liquidation_date,
        "data_recebimento": receipt_date,
    }
    memo[key] = result
    return result


def load_holiday_cache(cache_path: Path):
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_holiday_cache(cache_path: Path, cache: dict):
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_holidays_for_year(year: int) -> list:
    url = HOLIDAY_API_URL.format(year=year)
    response = requests.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    payload = response.json()

    dates = []
    for item in payload:
        dt = item.get("date")
        if dt:
            dates.append(dt)

    return sorted(set(dates))


def load_holidays(years, base_dir: Path) -> set:
    cache_path = base_dir / HOLIDAY_CACHE_FILENAME
    cache = load_holiday_cache(cache_path)
    holidays = set()

    for year in sorted(set(int(y) for y in years)):
        year_key = str(year)

        if year_key not in cache:
            try:
                print(f"Consultando feriados do ano {year} na API...")
                cache[year_key] = fetch_holidays_for_year(year)
            except Exception as exc:
                # Se a API de feriados estiver fora/sem internet, o processamento continua
                # usando apenas sábados e domingos como dias não úteis.
                # Caso já exista cache do ano, usa o cache normalmente.
                if year_key not in cache:
                    print(f"AVISO: não foi possível consultar feriados de {year}: {exc}")
                    cache[year_key] = []

        for iso_date in cache.get(year_key, []):
            holidays.add(pd.to_datetime(iso_date).date())

    for iso_date in ADDITIONAL_MANUAL_HOLIDAYS:
        holidays.add(pd.to_datetime(iso_date).date())

    save_holiday_cache(cache_path, cache)
    return holidays


def status_from_diff(diff: float) -> str:
    if pd.isna(diff):
        return ""
    if abs(float(diff)) < 0.005:
        return "OK"
    return "PGT MAIOR" if diff > 0 else "PGT MENOR"


def format_brl(value):
    if pd.isna(value):
        return "R$ 0,00"
    s = f"{float(value):,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def calcular_fluxo_recebimento_vetorizado(data_venda, prazo_base, prazo_delay, holidays: set):
    """Calcula início, liquidação e recebimento em lote.

    Essa função substitui o loop linha a linha. Na planilha enviada, depois de abrir
    crédito parcelado, existem centenas de milhares de previsões; loop Python fica
    muito lento. O numpy faz o mesmo cálculo de dia útil de forma vetorizada.
    """
    datas = pd.to_datetime(data_venda, errors="coerce").values.astype("datetime64[D]")
    base = np.asarray(prazo_base, dtype="timedelta64[D]")
    delay = np.asarray(prazo_delay, dtype="timedelta64[D]")

    holiday_values = []
    for h in sorted(holidays):
        try:
            holiday_values.append(np.datetime64(pd.Timestamp(h).date(), "D"))
        except Exception:
            pass
    holidays_np = np.array(holiday_values, dtype="datetime64[D]")

    # 1) Se a venda caiu em sábado/domingo/feriado, começa a contar no próximo útil.
    inicio = np.busday_offset(datas, 0, roll="forward", holidays=holidays_np)

    # 2) Soma prazo em dias corridos e empurra a liquidação para o próximo útil.
    liquidacao_candidata = inicio + base
    liquidacao = np.busday_offset(liquidacao_candidata, 0, roll="forward", holidays=holidays_np)

    # 3) Soma delay em dias corridos e empurra o recebimento para o próximo útil.
    recebimento_candidato = liquidacao + delay
    recebimento = np.busday_offset(recebimento_candidato, 0, roll="forward", holidays=holidays_np)

    inicio_ts = pd.to_datetime(inicio)
    liquidacao_ts = pd.to_datetime(liquidacao)
    recebimento_ts = pd.to_datetime(recebimento)

    return inicio_ts, liquidacao_ts, recebimento_ts


def prepare_base(df_base: pd.DataFrame, holidays: set):
    base = df_base.copy()
    base = base.dropna(how="all").reset_index(drop=True)

    col_data_venda = find_column(base, BASE_COL_ALIASES["data_venda"])
    col_valor_liquido = find_column(base, BASE_COL_ALIASES["valor_liquido"])
    col_valor_parcela = find_column(base, BASE_COL_ALIASES["valor_parcela"], required=False)
    col_produto = find_column(base, BASE_COL_ALIASES["produto"])
    col_parcelas = find_column(base, BASE_COL_ALIASES["parcelas"])
    col_documento = find_column(base, BASE_COL_ALIASES["documento"], required=False)
    col_bandeira = find_column(base, BASE_COL_ALIASES["bandeira"], required=False)

    base["LINHA_ORIGEM"] = np.arange(len(base)) + 2
    base["DATA_VENDA"] = parse_date_series(get_series(base, col_data_venda))
    base["VALOR_VENDA_LIQUIDO"] = parse_money_series(get_series(base, col_valor_liquido))
    base["VALOR_PARCELA_ORIGINAL"] = parse_money_series(get_series(base, col_valor_parcela)) if col_valor_parcela else np.nan
    base["PRODUTO"] = get_series(base, col_produto).astype(str).str.strip()
    base["PRODUTO_NORM"] = base["PRODUTO"].apply(normalize_text)
    base["TOTAL_PARCELAS"] = get_series(base, col_parcelas).apply(lambda x: parse_int_safe(x, 1))
    base["DOCUMENTO"] = get_series(base, col_documento).fillna("").astype(str).str.strip()
    base["BANDEIRA"] = get_series(base, col_bandeira).fillna("").astype(str).str.strip()

    invalid_mask = base["DATA_VENDA"].isna() | base["VALOR_VENDA_LIQUIDO"].isna() | (base["TOTAL_PARCELAS"] <= 0)
    descartadas = base[invalid_mask].copy()
    if not descartadas.empty:
        descartadas["MOTIVO_DESCARTE"] = np.select(
            [
                descartadas["DATA_VENDA"].isna(),
                descartadas["VALOR_VENDA_LIQUIDO"].isna(),
                descartadas["TOTAL_PARCELAS"] <= 0,
            ],
            ["DATA_VENDA_INVALIDA", "VALOR_LIQUIDO_INVALIDO", "N_DE_PARCELAS_INVALIDO"],
            default="LINHA_INVALIDA",
        )
    else:
        descartadas = pd.DataFrame(columns=list(base.columns) + ["MOTIVO_DESCARTE"])

    base_valid = base[~invalid_mask].copy()
    if base_valid.empty:
        validas = pd.DataFrame(columns=[
            "LINHA_ORIGEM", "DOCUMENTO", "BANDEIRA", "PRODUTO", "DATA_VENDA",
            "VALOR_VENDA_LIQUIDO", "TOTAL_PARCELAS", "PARCELA_ATUAL", "VALOR_PARCELA_ORIGINAL",
            "VALOR_LIQUIDO", "DATA_INICIO_CONTAGEM", "DATA_LIQUIDACAO",
            "PRAZO_BASE_DIAS_CORRIDOS", "PRAZO_DELAY_DIAS_CORRIDOS", "REGRA_APLICADA",
            "DATA_PREVISTA_RECEBIMENTO", "MES_PREVISTO", "DIA_SEMANA_PREVISTO",
            "DETALHE_CONCAT", "MOTIVO_DESCARTE"
        ])
        return validas, descartadas.reset_index(drop=True)

    base_valid["EH_CREDITO"] = base_valid["PRODUTO_NORM"].str.contains("CREDITO", na=False)
    base_valid["EH_DEBITO"] = base_valid["PRODUTO_NORM"].str.contains("DEBITO", na=False)
    base_valid["EH_PIX"] = base_valid["PRODUTO_NORM"].str.contains("PIX", na=False)
    base_valid["QTD_LINHAS_PREVISTAS"] = np.where(
        base_valid["EH_CREDITO"] & (base_valid["TOTAL_PARCELAS"] > 1),
        base_valid["TOTAL_PARCELAS"],
        1,
    ).astype(int)

    cols_expandir = [
        "LINHA_ORIGEM", "DOCUMENTO", "BANDEIRA", "PRODUTO", "PRODUTO_NORM",
        "DATA_VENDA", "VALOR_VENDA_LIQUIDO", "VALOR_PARCELA_ORIGINAL",
        "TOTAL_PARCELAS", "EH_CREDITO", "EH_DEBITO", "EH_PIX", "QTD_LINHAS_PREVISTAS"
    ]
    base_expandir = base_valid[cols_expandir].copy()
    expanded = base_expandir.loc[base_expandir.index.repeat(base_expandir["QTD_LINHAS_PREVISTAS"])].copy()
    expanded["PARCELA_ATUAL"] = expanded.groupby(level=0).cumcount() + 1
    expanded = expanded.reset_index(drop=True)

    expanded["VALOR_PARCELA_BASE"] = expanded["VALOR_PARCELA_ORIGINAL"]
    mask_sem_parcela = expanded["VALOR_PARCELA_BASE"].isna() | (expanded["VALOR_PARCELA_BASE"].astype(float) == 0)
    expanded.loc[mask_sem_parcela, "VALOR_PARCELA_BASE"] = (
        expanded.loc[mask_sem_parcela, "VALOR_VENDA_LIQUIDO"].astype(float)
        / expanded.loc[mask_sem_parcela, "TOTAL_PARCELAS"].astype(float)
    )

    parcelado_mask = expanded["EH_CREDITO"] & (expanded["TOTAL_PARCELAS"] > 1)
    expanded["VALOR_LIQUIDO"] = expanded["VALOR_VENDA_LIQUIDO"].astype(float).round(2)
    expanded.loc[parcelado_mask, "VALOR_LIQUIDO"] = expanded.loc[parcelado_mask, "VALOR_PARCELA_BASE"].astype(float).round(2)

    last_parcela_mask = parcelado_mask & (expanded["PARCELA_ATUAL"] == expanded["TOTAL_PARCELAS"])
    expanded.loc[last_parcela_mask, "VALOR_LIQUIDO"] = (
        expanded.loc[last_parcela_mask, "VALOR_VENDA_LIQUIDO"].astype(float)
        - (expanded.loc[last_parcela_mask, "VALOR_PARCELA_BASE"].astype(float).round(2)
           * (expanded.loc[last_parcela_mask, "TOTAL_PARCELAS"].astype(int) - 1))
    ).round(2)

    expanded["PRAZO_BASE_DIAS_CORRIDOS"] = np.nan
    expanded["PRAZO_DELAY_DIAS_CORRIDOS"] = np.nan
    expanded["REGRA_APLICADA"] = ""
    expanded["MOTIVO_DESCARTE"] = ""

    mask_debito = expanded["EH_DEBITO"]
    mask_pix = expanded["EH_PIX"]
    mask_credito = expanded["EH_CREDITO"]
    mask_credito_1x = mask_credito & (expanded["TOTAL_PARCELAS"] <= 1)
    mask_credito_parc = mask_credito & (expanded["TOTAL_PARCELAS"] > 1)
    mask_prod_invalido = ~(mask_debito | mask_pix | mask_credito)

    expanded.loc[mask_debito, "PRAZO_BASE_DIAS_CORRIDOS"] = 1
    expanded.loc[mask_debito, "PRAZO_DELAY_DIAS_CORRIDOS"] = 30
    expanded.loc[mask_debito, "REGRA_APLICADA"] = "DÉBITO = D+1 + 30 corridos → recebimento no próximo útil"

    expanded.loc[mask_pix, "PRAZO_BASE_DIAS_CORRIDOS"] = 1
    expanded.loc[mask_pix, "PRAZO_DELAY_DIAS_CORRIDOS"] = 0
    expanded.loc[mask_pix, "REGRA_APLICADA"] = "PIX = D+1 corrido → recebimento no próximo útil"

    expanded.loc[mask_credito_1x, "PRAZO_BASE_DIAS_CORRIDOS"] = 30
    expanded.loc[mask_credito_1x, "PRAZO_DELAY_DIAS_CORRIDOS"] = 30
    expanded.loc[mask_credito_1x, "REGRA_APLICADA"] = "CRÉDITO 1X = D+30 + 30 corridos → recebimento no próximo útil"

    expanded.loc[mask_credito_parc, "PRAZO_BASE_DIAS_CORRIDOS"] = expanded.loc[mask_credito_parc, "PARCELA_ATUAL"].astype(int) * 30
    expanded.loc[mask_credito_parc, "PRAZO_DELAY_DIAS_CORRIDOS"] = 30
    expanded.loc[mask_credito_parc, "REGRA_APLICADA"] = (
        "CRÉDITO PARCELADO "
        + expanded.loc[mask_credito_parc, "PARCELA_ATUAL"].astype(str)
        + "/"
        + expanded.loc[mask_credito_parc, "TOTAL_PARCELAS"].astype(str)
        + " = D+(parcela x 30) + 30 corridos → recebimento no próximo útil"
    )

    expanded.loc[mask_prod_invalido, "MOTIVO_DESCARTE"] = "Produto não suportado para regra de recebimento: " + expanded.loc[mask_prod_invalido, "PRODUTO"].astype(str)

    invalid_expanded = expanded[expanded["MOTIVO_DESCARTE"] != ""].copy()
    if not invalid_expanded.empty:
        descartadas = pd.concat([descartadas, invalid_expanded], ignore_index=True)

    validas = expanded[expanded["MOTIVO_DESCARTE"] == ""].copy()
    validas["PRAZO_BASE_DIAS_CORRIDOS"] = validas["PRAZO_BASE_DIAS_CORRIDOS"].astype(int)
    validas["PRAZO_DELAY_DIAS_CORRIDOS"] = validas["PRAZO_DELAY_DIAS_CORRIDOS"].astype(int)

    print("Calculando datas previstas de recebimento em modo rápido...")
    datas_inicio_validas, datas_liquidacao_validas, datas_recebimento_validas = calcular_fluxo_recebimento_vetorizado(
        validas["DATA_VENDA"],
        validas["PRAZO_BASE_DIAS_CORRIDOS"].astype(int).to_numpy(),
        validas["PRAZO_DELAY_DIAS_CORRIDOS"].astype(int).to_numpy(),
        holidays,
    )

    # Mantemos essas datas como texto ISO dentro da base expandida.
    # Isso reduz consumo de memória ao trabalhar com centenas de milhares de linhas.
    validas["DATA_INICIO_CONTAGEM"] = datas_inicio_validas.strftime("%Y-%m-%d")
    validas["DATA_LIQUIDACAO"] = datas_liquidacao_validas.strftime("%Y-%m-%d")
    validas["DATA_PREVISTA_RECEBIMENTO"] = datas_recebimento_validas.strftime("%Y-%m-%d")
    validas["MES_PREVISTO"] = datas_recebimento_validas.strftime("%m/%Y")
    validas["DIA_SEMANA_PREVISTO"] = datas_recebimento_validas.dayofweek.map(WEEKDAY_PT)


    validas["DETALHE_CONCAT"] = (
        "Doc " + validas["DOCUMENTO"].astype(str)
        + " - " + validas["PRODUTO"].astype(str)
        + " - Parcela " + validas["PARCELA_ATUAL"].astype(str)
        + "/" + validas["TOTAL_PARCELAS"].astype(str)
    )

    keep_cols = [
        "LINHA_ORIGEM", "DOCUMENTO", "BANDEIRA", "PRODUTO", "DATA_VENDA",
        "VALOR_VENDA_LIQUIDO", "TOTAL_PARCELAS", "PARCELA_ATUAL", "VALOR_PARCELA_ORIGINAL",
        "VALOR_LIQUIDO", "DATA_INICIO_CONTAGEM", "DATA_LIQUIDACAO",
        "PRAZO_BASE_DIAS_CORRIDOS", "PRAZO_DELAY_DIAS_CORRIDOS", "REGRA_APLICADA",
        "DATA_PREVISTA_RECEBIMENTO", "MES_PREVISTO", "DIA_SEMANA_PREVISTO",
        "DETALHE_CONCAT", "MOTIVO_DESCARTE"
    ]
    return validas[keep_cols].reset_index(drop=True), descartadas.reset_index(drop=True)

def prepare_extrato(df_extrato: pd.DataFrame):
    extr = df_extrato.copy()
    extr = extr.dropna(how="all").reset_index(drop=True)

    col_data = find_column(extr, EXTRATO_COL_ALIASES["data_extrato"])
    col_valor = find_column(extr, EXTRATO_COL_ALIASES["valor_extrato"], required=False)
    col_credito = find_column(extr, EXTRATO_COL_ALIASES["credito_extrato"], required=False)
    col_debito = find_column(extr, EXTRATO_COL_ALIASES["debito_extrato"], required=False)
    col_desc = find_column(extr, EXTRATO_COL_ALIASES["descricao"], required=False)

    extr["DATA_EXTRATO"] = parse_date_series(get_series(extr, col_data))

    if col_valor:
        extr["VALOR_EXTRATO"] = parse_money_series(get_series(extr, col_valor))
    else:
        if not col_credito:
            raise ValueError(
                "Não encontrei a coluna de valor do extrato. Ajuste EXTRATO_COL_ALIASES com os nomes reais da sua aba EXTRATO."
            )
        credito = parse_money_series(get_series(extr, col_credito)).fillna(0)
        debito = parse_money_series(get_series(extr, col_debito)).fillna(0) if col_debito else 0
        extr["VALOR_EXTRATO"] = credito - debito

    extr["DESCRICAO"] = get_series(extr, col_desc).fillna("").astype(str).str.strip()

    extr = extr[extr["DATA_EXTRATO"].notna()].copy()
    extr = extr[extr["VALOR_EXTRATO"].notna()].copy()

    if EXTRATO_ONLY_POSITIVE_VALUES:
        extr = extr[extr["VALOR_EXTRATO"] > 0].copy()

    extr = extr.reset_index(drop=True)
    extr["MES_EXTRATO"] = extr["DATA_EXTRATO"].dt.strftime("%m/%Y")
    extr["DIA_SEMANA_EXTRATO"] = extr["DATA_EXTRATO"].dt.dayofweek.map(WEEKDAY_PT)

    return extr


def build_previsto_concatenado(validas: pd.DataFrame):
    def detalhes_resumidos(s):
        itens = [str(x) for x in s if str(x).strip()]
        if len(itens) > 30:
            return " | ".join(itens[:30]) + f" | ... +{len(itens) - 30} lançamentos"
        return " | ".join(itens)

    agrupado = (
        validas.groupby("DATA_PREVISTA_RECEBIMENTO", as_index=False)
        .agg(
            MES=("MES_PREVISTO", "first"),
            DIA_DA_SEMANA=("DIA_SEMANA_PREVISTO", "first"),
            TOTAL_PREVISTO=("VALOR_LIQUIDO", "sum"),
            QTD_LANCAMENTOS=("VALOR_LIQUIDO", "size"),
            DETALHES=("DETALHE_CONCAT", detalhes_resumidos)
        )
        .sort_values("DATA_PREVISTA_RECEBIMENTO")
        .reset_index(drop=True)
    )
    agrupado["DATA_PREVISTA_RECEBIMENTO"] = pd.to_datetime(agrupado["DATA_PREVISTA_RECEBIMENTO"], errors="coerce")
    return agrupado


def build_extrato_concatenado(extr: pd.DataFrame):
    agrupado = (
        extr.groupby("DATA_EXTRATO", as_index=False)
        .agg(
            MES=("MES_EXTRATO", "first"),
            DIA_DA_SEMANA=("DIA_SEMANA_EXTRATO", "first"),
            TOTAL_EXTRATO=("VALOR_EXTRATO", "sum"),
            QTD_LANCAMENTOS=("VALOR_EXTRATO", "size"),
            DETALHES_EXTRATO=("DESCRICAO", lambda s: " | ".join([str(x) for x in list(s)[:30] if str(x).strip()]))
        )
        .sort_values("DATA_EXTRATO")
        .reset_index(drop=True)
    )
    return agrupado


def build_conciliacao(previsto_conc: pd.DataFrame, extrato_conc: pd.DataFrame):
    conciliacao = previsto_conc.merge(
        extrato_conc,
        left_on="DATA_PREVISTA_RECEBIMENTO",
        right_on="DATA_EXTRATO",
        how="outer"
    )

    conciliacao["DATA"] = conciliacao["DATA_PREVISTA_RECEBIMENTO"].combine_first(conciliacao["DATA_EXTRATO"])
    conciliacao["MES"] = conciliacao["MES_x"].combine_first(conciliacao["MES_y"])
    conciliacao["DIA_DA_SEMANA"] = conciliacao["DIA_DA_SEMANA_x"].combine_first(conciliacao["DIA_DA_SEMANA_y"])
    conciliacao["PROJETADO"] = conciliacao["TOTAL_PREVISTO"].fillna(0.0)
    conciliacao["VLR_EXTRATO"] = conciliacao["TOTAL_EXTRATO"].fillna(0.0)
    conciliacao["QTD_PREVISTA"] = conciliacao["QTD_LANCAMENTOS_x"].fillna(0).astype(int)
    conciliacao["QTD_EXTRATO"] = conciliacao["QTD_LANCAMENTOS_y"].fillna(0).astype(int)
    conciliacao["DIFERENCA"] = conciliacao["VLR_EXTRATO"] - conciliacao["PROJETADO"]
    conciliacao["STATUS"] = conciliacao["DIFERENCA"].apply(status_from_diff)
    conciliacao["DETALHES_PREVISTO"] = conciliacao["DETALHES"].fillna("")
    conciliacao["DETALHES_EXTRATO"] = conciliacao["DETALHES_EXTRATO"].fillna("")

    conciliacao = conciliacao[[
        "MES", "DATA", "PROJETADO", "VLR_EXTRATO", "DIFERENCA", "STATUS",
        "DIA_DA_SEMANA", "QTD_PREVISTA", "QTD_EXTRATO", "DETALHES_PREVISTO", "DETALHES_EXTRATO"
    ]].sort_values("DATA").reset_index(drop=True)

    return conciliacao


def build_calendar_daily(conciliacao: pd.DataFrame):
    min_date = conciliacao["DATA"].min()
    max_date = conciliacao["DATA"].max()

    if pd.isna(min_date) or pd.isna(max_date):
        raise ValueError("Não foi possível montar o calendário diário, pois não há datas válidas.")

    start = pd.Timestamp(year=min_date.year, month=min_date.month, day=1)
    end = pd.Timestamp(year=max_date.year, month=max_date.month, day=1) + pd.offsets.MonthEnd(1)

    calendario = pd.DataFrame({"DATA": pd.date_range(start, end, freq="D")})
    out = calendario.merge(conciliacao, on="DATA", how="left")

    out["MES"] = out["DATA"].dt.strftime("%m/%Y")
    out["DIA_DA_SEMANA"] = out["DATA"].dt.dayofweek.map(WEEKDAY_PT)
    out["PROJETADO"] = out["PROJETADO"].fillna(0.0)
    out["VLR_EXTRATO"] = out["VLR_EXTRATO"].fillna(0.0)
    
    out["DIFERENCA"] = out["VLR_EXTRATO"] - out["PROJETADO"]
    
    hoje = pd.Timestamp("today").normalize()
    out["STATUS"] = out.apply(
        lambda row: "A RECEBER" if row["DATA"] > hoje else status_from_diff(row["DIFERENCA"]),
        axis=1
    )
    
    out["QTD_PREVISTA"] = out["QTD_PREVISTA"].fillna(0).astype(int)
    out["QTD_EXTRATO"] = out["QTD_EXTRATO"].fillna(0).astype(int)
    out["DETALHES_PREVISTO"] = out["DETALHES_PREVISTO"].fillna("")
    out["DETALHES_EXTRATO"] = out["DETALHES_EXTRATO"].fillna("")

    return out


def build_monthly(calendar_daily: pd.DataFrame, extr: pd.DataFrame | None = None):
    # Para o último mês existente no extrato, compara somente até a última data real
    # do extrato. Isso evita comparar junho inteiro contra extrato lançado só até 01/06, por exemplo.
    max_extrato_date = None
    if extr is not None and not extr.empty and "DATA_EXTRATO" in extr.columns:
        max_extrato_date = extr["DATA_EXTRATO"].max()

    rows = []
    for period, block in calendar_daily.groupby(calendar_daily["DATA"].dt.to_period("M")):
        block_calc = block.copy()
        status_forcado = None

        if max_extrato_date is not None:
            max_period = max_extrato_date.to_period("M")
            if period == max_period:
                block_calc = block_calc[block_calc["DATA"] <= max_extrato_date]
            elif period > max_period:
                status_forcado = "A RECEBER"

        projetado = float(block_calc["PROJETADO"].sum())
        vlr_extrato = float(block_calc["VLR_EXTRATO"].sum())
        diferenca = vlr_extrato - projetado
        status = status_forcado if status_forcado else status_from_diff(diferenca)

        rows.append({
            "MES_REF": period,
            "MES": period.strftime("%m/%Y"),
            "PROJETADO": projetado,
            "VLR_EXTRATO": vlr_extrato,
            "DIFERENCA": diferenca,
            "STATUS": status,
        })

    mensal = pd.DataFrame(rows)
    if mensal.empty:
        return pd.DataFrame(columns=["MES_REF", "MES", "PROJETADO", "VLR_EXTRATO", "DIFERENCA", "STATUS"])
    return mensal[["MES_REF", "MES", "PROJETADO", "VLR_EXTRATO", "DIFERENCA", "STATUS"]]

def build_resumo_final(calendar_daily: pd.DataFrame):
    resumo = calendar_daily[["DATA", "PROJETADO", "VLR_EXTRATO", "DIFERENCA", "STATUS"]].copy()
    resumo.columns = ["DATA", "VALOR PREVISÃO", "VALOR ENTRADA", "DIFERENÇA", "STATUS"]
    resumo = resumo[(resumo["VALOR PREVISÃO"] != 0) | (resumo["VALOR ENTRADA"] != 0)].reset_index(drop=True)
    return resumo


def build_resumo_executivo(calendar_daily: pd.DataFrame, validas: pd.DataFrame, extr: pd.DataFrame):
    total_previsto = float(calendar_daily["PROJETADO"].sum())
    total_entrada = float(calendar_daily["VLR_EXTRATO"].sum())
    total_diferenca = float(calendar_daily["DIFERENCA"].sum())

    dias_ok = int((calendar_daily["STATUS"] == "OK").sum())
    dias_maior = int((calendar_daily["STATUS"] == "PGT MAIOR").sum())
    dias_menor = int((calendar_daily["STATUS"] == "PGT MENOR").sum())

    qtd_previstas = int(len(validas))
    qtd_extrato = int(len(extr))

    return pd.DataFrame({
        "INDICADOR": [
            "TOTAL PREVISTO",
            "TOTAL ENTRADA",
            "DIFERENÇA TOTAL",
            "QTDE VENDAS PROCESSADAS",
            "QTDE LANÇAMENTOS EXTRATO",
            "DIAS COM STATUS OK",
            "DIAS COM PGT MAIOR",
            "DIAS COM PGT MENOR",
        ],
        "VALOR": [
            total_previsto,
            total_entrada,
            total_diferenca,
            qtd_previstas,
            qtd_extrato,
            dias_ok,
            dias_maior,
            dias_menor,
        ]
    })


def autosize_sheet(ws, extra=2):
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(max_length + extra, 60)


def _header_to_index(ws):
    return {normalize_text(cell.value): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value is not None}


def style_basic_sheet(ws, currency_headers=None, date_headers=None, wrap_headers=None):
    currency_headers = currency_headers or []
    date_headers = date_headers or []
    wrap_headers = wrap_headers or []

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    header_map = _header_to_index(ws)
    currency_indexes = {header_map.get(normalize_text(h)) for h in currency_headers}
    date_indexes = {header_map.get(normalize_text(h)) for h in date_headers}
    wrap_indexes = {header_map.get(normalize_text(h)) for h in wrap_headers}
    currency_indexes.discard(None)
    date_indexes.discard(None)
    wrap_indexes.discard(None)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column in currency_indexes and isinstance(cell.value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
            if cell.column in date_indexes and cell.value is not None:
                cell.number_format = 'dd/mm/yyyy'
            if cell.column in wrap_indexes:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    autosize_sheet(ws)


def style_resumo_executivo(ws):
    header_fill = PatternFill("solid", fgColor="385D9D")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column == 2 and isinstance(cell.value, (int, float)):
                if row[0].value in ["TOTAL PREVISTO", "TOTAL ENTRADA", "DIFERENÇA TOTAL"]:
                    cell.number_format = 'R$ #,##0.00'

    ws.freeze_panes = "A2"
    autosize_sheet(ws, extra=4)


def write_panel(ws, daily_calendar: pd.DataFrame, mensal: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_fill_blue = PatternFill("solid", fgColor="385D9D")
    title_fill_beige = PatternFill("solid", fgColor="EAD7A4")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="1F1F1F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    start_row_summary = 2
    start_col_summary = 1

    ws.merge_cells(
        start_row=start_row_summary,
        start_column=start_col_summary,
        end_row=start_row_summary,
        end_column=start_col_summary + 4,
    )
    c = ws.cell(start_row_summary, start_col_summary, "CONSOLIDADO POR MÊS")
    c.fill = title_fill_blue
    c.font = white_bold
    c.alignment = center

    headers_summary = ["DIA/MÊS", "PROJETADO", "VLR EXTRATO", "DIFERENÇA", "STATUS"]
    for j, header in enumerate(headers_summary, start=start_col_summary):
        cell = ws.cell(start_row_summary + 2, j, header)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border

    row = start_row_summary + 3
    for _, r in mensal.iterrows():
        values = [r["MES"], r["PROJETADO"], r["VLR_EXTRATO"], r["DIFERENCA"], r["STATUS"]]
        for j, value in enumerate(values, start=start_col_summary):
            cell = ws.cell(row, j, value)
            cell.border = border
            if j in (2, 3, 4):
                cell.number_format = 'R$ #,##0.00'
            if j == 5:
                if value == "PGT MAIOR":
                    cell.font = Font(color="008000", bold=True)
                elif value == "PGT MENOR":
                    cell.font = Font(color="C00000", bold=True)
                elif value == "A RECEBER":
                    cell.font = Font(color="808080", bold=True)
        row += 1

    total_values = [
        "TOTAL",
        float(mensal["PROJETADO"].sum()),
        float(mensal["VLR_EXTRATO"].sum()),
        float(mensal["DIFERENCA"].sum()),
        status_from_diff(float(mensal["DIFERENCA"].sum())),
    ]
    for j, value in enumerate(total_values, start=start_col_summary):
        cell = ws.cell(row, j, value)
        cell.border = border
        cell.font = Font(bold=True)
        if j in (2, 3, 4):
            cell.number_format = 'R$ #,##0.00'

    months = list(daily_calendar["DATA"].dt.to_period("M").drop_duplicates().sort_values())
    base_col = 7
    current_top_row = 2
    block_gap_cols = 1
    block_width = 6
    max_block_height_in_band = 0

    for idx, period in enumerate(months):
        position_in_band = idx % 2

        if idx > 0 and position_in_band == 0:
            current_top_row += max_block_height_in_band + 3
            max_block_height_in_band = 0

        start_col = base_col + position_in_band * (block_width + block_gap_cols)
        block = daily_calendar[daily_calendar["DATA"].dt.to_period("M") == period].copy()

        ws.merge_cells(
            start_row=current_top_row,
            start_column=start_col,
            end_row=current_top_row,
            end_column=start_col + block_width - 1,
        )
        tc = ws.cell(current_top_row, start_col, "CONSOLIDADO POR DIA E MÊS")
        tc.fill = title_fill_beige
        tc.font = Font(bold=True)
        tc.alignment = center

        headers = ["MÊS", "DATA", "PROJETADO", "VLR EXTRATO", "DIFERENÇA", "DIA DA SEMANA"]
        for j, header in enumerate(headers, start=start_col):
            hc = ws.cell(current_top_row + 2, j, header)
            hc.fill = header_fill
            hc.font = Font(bold=True)
            hc.alignment = center
            hc.border = border

        write_row = current_top_row + 3
        for _, r in block.iterrows():
            vals = [
                r["MES"],
                r["DATA"].to_pydatetime(),
                float(r["PROJETADO"]),
                float(r["VLR_EXTRATO"]),
                float(r["DIFERENCA"]),
                r["DIA_DA_SEMANA"],
            ]
            for j, value in enumerate(vals, start=start_col):
                cell = ws.cell(write_row, j, value)
                cell.border = border
                if j == start_col + 1:
                    cell.number_format = "dd/mm/yyyy"
                if j in (start_col + 2, start_col + 3, start_col + 4):
                    cell.number_format = 'R$ #,##0.00'
            write_row += 1

        totals = [
            "",
            "TOTAL",
            float(block["PROJETADO"].sum()),
            float(block["VLR_EXTRATO"].sum()),
            float(block["DIFERENCA"].sum()),
            "",
        ]
        for j, value in enumerate(totals, start=start_col):
            cell = ws.cell(write_row, j, value)
            cell.border = border
            cell.font = Font(bold=True)
            if j in (start_col + 2, start_col + 3, start_col + 4):
                cell.number_format = 'R$ #,##0.00'

        block_height = len(block) + 4
        max_block_height_in_band = max(max_block_height_in_band, block_height)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["E"].width = 14


def limitar_abas_detalhadas(df: pd.DataFrame, nome: str) -> pd.DataFrame:
    if EXPORTAR_ABAS_DETALHADAS_COMPLETAS:
        return df.copy()

    total = len(df)
    if total <= LIMITE_LINHAS_ABAS_DETALHADAS:
        return df.copy()

    print(
        f"AVISO: aba {nome} possui {total:,} linhas. "
        f"Exportando amostra de {LIMITE_LINHAS_ABAS_DETALHADAS:,} linhas para não travar o Excel."
        .replace(",", ".")
    )
    return df.head(LIMITE_LINHAS_ABAS_DETALHADAS).copy()


def write_output(
    output_path: Path,
    input_path: Path,
    base_sheet_name: str,
    extrato_sheet_name: str,
    validas: pd.DataFrame,
    descartadas: pd.DataFrame,
    extr: pd.DataFrame,
    previsto_conc: pd.DataFrame,
    extrato_conc: pd.DataFrame,
    conciliacao: pd.DataFrame,
    daily_calendar: pd.DataFrame,
    mensal: pd.DataFrame,
):
    print("Preparando abas do Excel...")

    validas_detalhe = limitar_abas_detalhadas(validas, "PREVISTO_DETALHADO")
    validas_auditoria = limitar_abas_detalhadas(validas, "AUDITORIA_REGRAS")

    detalhado_previsto = validas_detalhe[[
        "LINHA_ORIGEM", "DOCUMENTO", "BANDEIRA", "DATA_VENDA", "DATA_INICIO_CONTAGEM", "DATA_LIQUIDACAO",
        "PRODUTO", "TOTAL_PARCELAS", "PARCELA_ATUAL", "VALOR_VENDA_LIQUIDO", "VALOR_PARCELA_ORIGINAL",
        "VALOR_LIQUIDO", "PRAZO_BASE_DIAS_CORRIDOS", "PRAZO_DELAY_DIAS_CORRIDOS",
        "REGRA_APLICADA", "DATA_PREVISTA_RECEBIMENTO", "MES_PREVISTO",
        "DIA_SEMANA_PREVISTO", "DETALHE_CONCAT"
    ]].copy().rename(columns={
        "LINHA_ORIGEM": "LINHA ORIGEM",
        "DATA_VENDA": "DATA DA VENDA",
        "DATA_INICIO_CONTAGEM": "DATA INÍCIO CONTAGEM",
        "DATA_LIQUIDACAO": "DATA LIQUIDAÇÃO",
        "TOTAL_PARCELAS": "TOTAL DE PARCELAS",
        "PARCELA_ATUAL": "PARCELA ATUAL",
        "VALOR_VENDA_LIQUIDO": "VALOR LÍQUIDO DA VENDA",
        "VALOR_PARCELA_ORIGINAL": "VLR PARCELA ORIGINAL",
        "VALOR_LIQUIDO": "VALOR PREVISTO DA PARCELA",
        "PRAZO_BASE_DIAS_CORRIDOS": "PRAZO BASE DIAS CORRIDOS",
        "PRAZO_DELAY_DIAS_CORRIDOS": "PRAZO DELAY DIAS CORRIDOS",
        "REGRA_APLICADA": "REGRA APLICADA",
        "DATA_PREVISTA_RECEBIMENTO": "DATA PREVISTA RECEBIMENTO",
        "MES_PREVISTO": "MÊS PREVISTO",
        "DIA_SEMANA_PREVISTO": "DIA DA SEMANA PREVISTO",
        "DETALHE_CONCAT": "DETALHE CONCATENADO",
    })

    previsto_concat = previsto_conc.rename(columns={
        "DATA_PREVISTA_RECEBIMENTO": "DATA PREVISTA",
        "MES": "MÊS",
        "DIA_DA_SEMANA": "DIA DA SEMANA",
        "TOTAL_PREVISTO": "TOTAL PREVISTO",
        "QTD_LANCAMENTOS": "QTD LANÇAMENTOS",
        "DETALHES": "DETALHES",
    })

    extrato_tratado = extr[["DATA_EXTRATO", "VALOR_EXTRATO", "MES_EXTRATO", "DIA_SEMANA_EXTRATO", "DESCRICAO"]].copy().rename(columns={
        "DATA_EXTRATO": "DATA EXTRATO",
        "VALOR_EXTRATO": "VALOR EXTRATO",
        "MES_EXTRATO": "MÊS EXTRATO",
        "DIA_SEMANA_EXTRATO": "DIA DA SEMANA EXTRATO",
        "DESCRICAO": "DESCRIÇÃO",
    })

    extrato_concat_export = extrato_conc.rename(columns={
        "DATA_EXTRATO": "DATA EXTRATO",
        "MES": "MÊS",
        "DIA_DA_SEMANA": "DIA DA SEMANA",
        "TOTAL_EXTRATO": "TOTAL EXTRATO",
        "QTD_LANCAMENTOS": "QTD LANÇAMENTOS",
        "DETALHES_EXTRATO": "DETALHES EXTRATO",
    })

    conciliacao_export = conciliacao.rename(columns={
        "MES": "MÊS",
        "DATA": "DATA",
        "PROJETADO": "PROJETADO",
        "VLR_EXTRATO": "VLR EXTRATO",
        "DIFERENCA": "DIFERENÇA",
        "STATUS": "STATUS",
        "DIA_DA_SEMANA": "DIA DA SEMANA",
        "QTD_PREVISTA": "QTD PREVISTA",
        "QTD_EXTRATO": "QTD EXTRATO",
        "DETALHES_PREVISTO": "DETALHES PREVISTO",
        "DETALHES_EXTRATO": "DETALHES EXTRATO",
    })

    resumo_final = build_resumo_final(daily_calendar)
    resumo_executivo = build_resumo_executivo(daily_calendar, validas, extr)

    auditoria = validas_auditoria[[
        "LINHA_ORIGEM", "DOCUMENTO", "BANDEIRA", "DATA_VENDA", "DATA_INICIO_CONTAGEM", "DATA_LIQUIDACAO",
        "PRODUTO", "TOTAL_PARCELAS", "PARCELA_ATUAL", "VALOR_VENDA_LIQUIDO", "VALOR_LIQUIDO",
        "PRAZO_BASE_DIAS_CORRIDOS", "PRAZO_DELAY_DIAS_CORRIDOS", "REGRA_APLICADA", "DATA_PREVISTA_RECEBIMENTO"
    ]].copy().rename(columns={
        "LINHA_ORIGEM": "LINHA ORIGEM",
        "DATA_VENDA": "DATA DA VENDA",
        "DATA_INICIO_CONTAGEM": "DATA INÍCIO CONTAGEM",
        "DATA_LIQUIDACAO": "DATA LIQUIDAÇÃO",
        "TOTAL_PARCELAS": "TOTAL DE PARCELAS",
        "PARCELA_ATUAL": "PARCELA ATUAL",
        "VALOR_VENDA_LIQUIDO": "VALOR LÍQUIDO DA VENDA",
        "VALOR_LIQUIDO": "VALOR PREVISTO DA PARCELA",
        "PRAZO_BASE_DIAS_CORRIDOS": "PRAZO BASE DIAS CORRIDOS",
        "PRAZO_DELAY_DIAS_CORRIDOS": "PRAZO DELAY DIAS CORRIDOS",
        "REGRA_APLICADA": "REGRA APLICADA",
        "DATA_PREVISTA_RECEBIMENTO": "DATA PREVISTA",
    })

    mensal_export = mensal[["MES", "PROJETADO", "VLR_EXTRATO", "DIFERENCA", "STATUS"]].copy().rename(columns={
        "MES": "DIA/MÊS",
        "PROJETADO": "PROJETADO",
        "VLR_EXTRATO": "VLR EXTRATO",
        "DIFERENCA": "DIFERENÇA",
        "STATUS": "STATUS",
    })

    modo_detalhe = (
        "COMPLETO" if EXPORTAR_ABAS_DETALHADAS_COMPLETAS
        else f"AMOSTRA DE {LIMITE_LINHAS_ABAS_DETALHADAS} LINHAS"
    )

    parametros = pd.DataFrame({
        "PARÂMETRO": [
            "ARQUIVO DE ENTRADA",
            "ABA BASE",
            "ABA EXTRATO",
            "API FERIADOS",
            "EXTENSÃO LIDA",
            "SOMENTE VALORES POSITIVOS NO EXTRATO",
            "FERIADOS MANUAIS ADICIONAIS",
            "REGRA DÉBITO",
            "REGRA PIX",
            "REGRA CRÉDITO 1X",
            "REGRA CRÉDITO PARCELADO",
            "LINHAS PREVISTAS GERADAS",
            "MODO DAS ABAS DETALHADAS",
            "OBSERVAÇÃO GERAL",
        ],
        "VALOR": [
            input_path.name,
            base_sheet_name,
            extrato_sheet_name,
            HOLIDAY_API_URL,
            input_path.suffix.lower(),
            str(EXTRATO_ONLY_POSITIVE_VALUES),
            ", ".join(ADDITIONAL_MANUAL_HOLIDAYS) if ADDITIONAL_MANUAL_HOLIDAYS else "",
            "Venda em dia não útil começa no próximo útil; D+1 + 30 corridos; recebimento no próximo útil",
            "Venda em dia não útil começa no próximo útil; D+1 corrido; recebimento no próximo útil",
            "Venda em dia não útil começa no próximo útil; D+30 + 30 corridos; recebimento no próximo útil",
            "A venda é aberta em parcelas; cada parcela usa VLR PARCELA; parcela N recebe em D+(N x 30) + 30 corridos; recebimento no próximo útil",
            len(validas),
            modo_detalhe,
            "Crédito parcelado agora é expandido em uma linha por parcela. O último mês do extrato é comparado somente até a última data real do extrato. Para evitar travamento, as abas PREVISTO_DETALHADO e AUDITORIA_REGRAS saem como amostra por padrão.",
        ],
    })

    print("Gravando arquivo Excel em modo rápido...")
    painel_export = mensal_export.copy()

    # IMPORTANTE:
    # Não usamos load_workbook/style em cima do arquivo inteiro neste modo,
    # porque o openpyxl pode demorar muitos minutos e parecer travado.
    # O arquivo sai leve, abre no Excel e traz todos os consolidados principais.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        painel_export.to_excel(writer, sheet_name="PAINEL", index=False)
        resumo_executivo.to_excel(writer, sheet_name="RESUMO_EXECUTIVO", index=False)
        resumo_final.to_excel(writer, sheet_name="RESUMO_FINAL", index=False)
        mensal_export.to_excel(writer, sheet_name="CONSOLIDADO_MES", index=False)
        conciliacao_export.to_excel(writer, sheet_name="CONCILIACAO_DIA", index=False)
        previsto_concat.to_excel(writer, sheet_name="PREVISTO_CONCATENADO", index=False)
        extrato_tratado.to_excel(writer, sheet_name="EXTRATO_TRATADO", index=False)
        extrato_concat_export.to_excel(writer, sheet_name="EXTRATO_CONCATENADO", index=False)
        detalhado_previsto.to_excel(writer, sheet_name="PREVISTO_DETALHADO", index=False)
        auditoria.to_excel(writer, sheet_name="AUDITORIA_REGRAS", index=False)
        descartadas.to_excel(writer, sheet_name="LINHAS_DESCARTADAS", index=False)
        parametros.to_excel(writer, sheet_name="PARAMETROS", index=False)

    print("Arquivo Excel gravado.")

def main():
    base_dir = get_app_dir()
    output_path = base_dir / OUTPUT_FILENAME

    print("=" * 80)
    print("INICIANDO PROCESSAMENTO DE RECEBIMENTOS STONE")
    print(f"Pasta de trabalho: {base_dir}")

    input_path = detect_input_file(base_dir)
    print(f"Arquivo encontrado: {input_path.name}")

    engine = get_engine(input_path)
    xls = pd.ExcelFile(input_path, engine=engine)

    sheet_base = resolve_sheet_name(xls, SHEET_BASE_CANDIDATES)
    sheet_extrato = resolve_sheet_name(xls, SHEET_EXTRATO_CANDIDATES)

    print(f"Aba base localizada: {sheet_base}")
    print(f"Aba extrato localizada: {sheet_extrato}")

    df_base = pd.read_excel(input_path, sheet_name=sheet_base, engine=engine, dtype=object)
    df_extrato = pd.read_excel(input_path, sheet_name=sheet_extrato, engine=engine, dtype=object)

    col_data_venda = find_column(df_base, BASE_COL_ALIASES["data_venda"])
    anos_base = parse_date_series(get_series(df_base, col_data_venda)).dropna().dt.year.unique().tolist()

    if not anos_base:
        raise ValueError("Não foi possível identificar anos válidos na coluna de DATA DA VENDA.")

    anos_feriados = set(anos_base)
    anos_feriados.update(y + 1 for y in anos_base)
    anos_feriados.update(y + 2 for y in anos_base)

    holidays = load_holidays(anos_feriados, base_dir)
    print(f"Total de feriados carregados: {len(holidays)}")

    validas, descartadas = prepare_base(df_base, holidays)
    print(f"Linhas válidas na base: {len(validas)}")
    print(f"Linhas descartadas na base: {len(descartadas)}")

    extr = prepare_extrato(df_extrato)
    print(f"Linhas válidas no extrato: {len(extr)}")

    previsto_conc = build_previsto_concatenado(validas)
    extrato_conc = build_extrato_concatenado(extr)
    conciliacao = build_conciliacao(previsto_conc, extrato_conc)
    daily_calendar = build_calendar_daily(conciliacao)
    mensal = build_monthly(daily_calendar, extr)

    print(f"Período consolidado: {daily_calendar['DATA'].min().date()} até {daily_calendar['DATA'].max().date()}")

    write_output(
        output_path=output_path,
        input_path=input_path,
        base_sheet_name=sheet_base,
        extrato_sheet_name=sheet_extrato,
        validas=validas,
        descartadas=descartadas,
        extr=extr,
        previsto_conc=previsto_conc,
        extrato_conc=extrato_conc,
        conciliacao=conciliacao,
        daily_calendar=daily_calendar,
        mensal=mensal,
    )

    print(f"Arquivo final gerado com sucesso: {output_path}")
    print("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\nERRO DURANTE O PROCESSAMENTO:")
        print(str(e))
        print("\nDETALHES:")
        print(traceback.format_exc())
        try:
            input("\nPressione ENTER para sair...")
        except Exception:
            pass
        raise